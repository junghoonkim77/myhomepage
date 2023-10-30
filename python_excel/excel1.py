import openpyxl

#wb = openpyxl.Workbook()

# ws =wb.active
# ws['a1']=123
# wb.save('myhomepage\python_excel\테스트.xlsx')

wb = openpyxl.load_workbook('myhomepage\python_excel\테스트.xlsx')
ws = wb.active
print(ws['a1'].value)

